"""Сортировка товаров в отгрузке МойСклад по ячейкам склада.

Использование:
    python sort_demand.py                          # стандартный запуск
    python sort_demand.py --days 14               # за 14 дней
    python sort_demand.py --apply                 # + сохранить порядок в МойСклад
    python sort_demand.py --state-name "сборка"  # другое название статуса
    python sort_demand.py --cell-attr "Место"    # другое название атрибута ячейки
"""

from __future__ import annotations

import collections
import dataclasses
import datetime
import pathlib
import json
import re
import sys
import threading
import time

import qrcode as _qrcode
import keyring
import openpyxl
import openpyxl.styles
import requests
import tyro
from colorama import Fore, Style, init

init(autoreset=True)

__version__ = "1.3.0"

BASE_URL = "https://api.moysklad.ru/api/remap/1.2"

# Фразы в комментарии, сигнализирующие что документ уже обработан ТСД Клеверенс
_CLEVERENCE_MARKERS = (
    "document has been picked on the cleverence handheld terminal",
    "документ отобран на тсд клеверенс",
)


# ── Config ────────────────────────────────────────────────────────────────────


@dataclasses.dataclass
class AppConfig:
    """Параметры сортировки отгрузок МойСклад."""

    days: int = 30
    """Количество дней назад для выборки отгрузок."""

    state_name: str = "на сборке"
    """Название статуса для фильтрации отгрузок."""

    cell_attr: str = "Ячейка"
    """Название пользовательского атрибута ячейки у товара."""

    apply: bool = False
    """Сохранить отсортированный порядок позиций в МойСклад."""

    debug: bool = False
    """Выводить сырой ответ API для диагностики."""


# ── Rate Limiter ──────────────────────────────────────────────────────────────

# МойСклад: Bearer-токен → 45 запросов за 3-секундное окно
# Берём 40 для запаса; при 429 ждём Retry-After (или экспоненциальный backoff)
_RATE_LIMIT_REQUESTS = 40
_RATE_LIMIT_WINDOW = 3.0  # секунды
_MAX_RETRIES = 5


class _RateLimiter:
    """Thread-safe token bucket rate limiter."""

    def __init__(self, rate: int, window: float) -> None:
        self._rate = rate          # токенов на окно
        self._window = window      # размер окна в секундах
        self._tokens = float(rate)
        self._last = time.monotonic()
        self._lock = threading.Lock()

    def acquire(self) -> None:
        """Блокирует вызывающий поток до получения разрешения на запрос."""
        with self._lock:
            now = time.monotonic()
            elapsed = now - self._last
            # Пополняем токены пропорционально прошедшему времени
            self._tokens = min(
                float(self._rate),
                self._tokens + elapsed * self._rate / self._window,
            )
            self._last = now

            if self._tokens < 1.0:
                wait = (1.0 - self._tokens) * self._window / self._rate
                time.sleep(wait)
                self._tokens = 0.0
            else:
                self._tokens -= 1.0


_rate_limiter = _RateLimiter(_RATE_LIMIT_REQUESTS, _RATE_LIMIT_WINDOW)


# ── API Client ────────────────────────────────────────────────────────────────


class MoySkladClient:
    def __init__(self, token: str) -> None:
        self._session = requests.Session()
        self._session.headers.update(
            {
                "Authorization": f"Bearer {token}",
                "Accept-Encoding": "gzip",
                "Content-Type": "application/json",
            }
        )

    # ── internal ──────────────────────────────────────────────────────────────

    def _request(
        self,
        method: str,
        url: str,
        params: dict | None = None,
        json: dict | list | None = None,
    ) -> dict | list:
        """Выполнить запрос с rate limiting и retry при 429/5xx."""
        for attempt in range(1, _MAX_RETRIES + 1):
            _rate_limiter.acquire()
            try:
                resp = self._session.request(
                    method, url, params=params, json=json, timeout=30
                )
            except requests.Timeout:
                wait = 2 ** attempt
                print(
                    f"{Fore.YELLOW}[WARN] Timeout, повтор через {wait}с "
                    f"(попытка {attempt}/{_MAX_RETRIES})…{Style.RESET_ALL}"
                )
                time.sleep(wait)
                continue

            if resp.status_code == 429:
                retry_after = float(resp.headers.get("Retry-After", 2 ** attempt))
                print(
                    f"{Fore.YELLOW}[WARN] 429 Rate limit, ожидание {retry_after:.1f}с "
                    f"(попытка {attempt}/{_MAX_RETRIES})…{Style.RESET_ALL}"
                )
                time.sleep(retry_after)
                continue

            if resp.status_code >= 500 and attempt < _MAX_RETRIES:
                wait = 2 ** attempt
                print(
                    f"{Fore.YELLOW}[WARN] HTTP {resp.status_code}, повтор через {wait}с "
                    f"(попытка {attempt}/{_MAX_RETRIES})…{Style.RESET_ALL}"
                )
                time.sleep(wait)
                continue

            if not resp.ok:
                try:
                    detail = resp.json()
                except Exception:
                    detail = resp.text
                raise RuntimeError(
                    f"HTTP {resp.status_code}\n"
                    f"  request.url = {resp.request.url}\n"
                    f"  resp.url    = {resp.url}\n"
                    f"{detail}"
                )

            # 204 No Content — некоторые bulk-операции не возвращают тело
            if not resp.content:
                return {}
            return resp.json()

        raise RuntimeError(f"Превышено {_MAX_RETRIES} попыток для {url}")

    # ── public ────────────────────────────────────────────────────────────────

    def get(self, path: str, params: dict | None = None) -> dict:
        return self._request("GET", f"{BASE_URL}{path}", params=params)

    def get_by_href(self, href: str, params: dict | None = None) -> dict:
        return self._request("GET", href, params=params)

    def post(self, path: str, payload: dict | list) -> dict | list:
        return self._request("POST", f"{BASE_URL}{path}", json=payload)

    def put(self, path: str, payload: dict) -> dict:
        return self._request("PUT", f"{BASE_URL}{path}", json=payload)


# ── Auth ──────────────────────────────────────────────────────────────────────


def load_token() -> str:
    token = keyring.get_password("moysklad", "access_token")
    if not token:
        _err(
            "Токен не найден. Запустите get_token.py для авторизации.",
            exit_code=1,
        )
    return token  # type: ignore[return-value]


# ── Business Logic ────────────────────────────────────────────────────────────


def find_state_href(client: MoySkladClient, state_name: str) -> str:
    """Найти href статуса по имени в метаданных отгрузок."""
    meta = client.get("/entity/demand/metadata")
    states: list[dict] = meta.get("states", [])
    for state in states:
        if state["name"].lower() == state_name.lower():
            return state["meta"]["href"]

    available = ", ".join(f"'{s['name']}'" for s in states)
    _err(
        f"Статус '{state_name}' не найден. Доступные статусы: {available}",
        exit_code=1,
    )
    return ""  # unreachable


def change_demand_state(client: MoySkladClient, demand_id: str, state_href: str) -> None:
    """Сменить статус отгрузки по UUID."""
    client.put(
        f"/entity/demand/{demand_id}",
        {"state": {"meta": {"href": state_href, "type": "state", "mediaType": "application/json"}}},
    )


def _is_cleverence_processed(demand: dict) -> bool:
    """True если в комментарии отгрузки есть маркер Клеверенс (обработан на ТСД)."""
    description: str = (demand.get("description") or "").lower()
    return any(marker in description for marker in _CLEVERENCE_MARKERS)


def fetch_demands(
    client: MoySkladClient,
    since: datetime.datetime,
    state_href: str,
) -> list[dict]:
    """Получить отгрузки за период с нужным статусом (все страницы).

    Документы с маркером Клеверенс в комментарии автоматически исключаются.
    """
    moment_str = since.strftime("%Y-%m-%d %H:%M:%S")
    rows: list[dict] = []
    offset = 0
    limit = 100

    while True:
        data = client.get(
            "/entity/demand",
            params={
                "filter": f"moment>{moment_str};state={state_href}",
                "limit": limit,
                "offset": offset,
                "expand": "state,agent",
                "order": "moment,desc",
            },
        )
        page = data.get("rows", [])
        rows.extend(page)
        if len(page) < limit:
            break
        offset += limit

    # Фильтр: исключаем документы уже обработанные ТСД Клеверенс
    filtered = [d for d in rows if not _is_cleverence_processed(d)]
    skipped = len(rows) - len(filtered)
    if skipped:
        print(
            f"{Fore.YELLOW}⚠ Пропущено {skipped} док. с маркером Клеверенс (уже обработаны на ТСД).{Style.RESET_ALL}"
        )
    return filtered


def fetch_positions(
    client: MoySkladClient,
    demand_id: str,
    cell_attr: str,
    debug: bool = False,
) -> list[dict]:
    """Получить позиции отгрузки со всеми страницами, каждой назначить ячейку.

    Ячейка берётся из поля `slot` позиции (место хранения МойСклад).
    Если slot отсутствует — fallback на пользовательский атрибут товара.
    """
    # Шаг 1: href позиций из самого документа
    demand_data = client.get(f"/entity/demand/{demand_id}")
    positions_meta: dict = demand_data.get("positions", {}).get("meta", {})
    positions_href: str = positions_meta.get("href", "")
    total_size: int = positions_meta.get("size", 0)

    if debug:
        print(
            f"{Fore.YELLOW}[DEBUG] positions href: {positions_href}"
            f"\n[DEBUG] total size: {total_size}{Style.RESET_ALL}"
        )

    if not positions_href:
        print(f"{Fore.RED}[WARN] Не удалось определить href позиций.{Style.RESET_ALL}")
        return []

    # Шаг 2: пагинация, раскрываем и assortment, и slot
    positions: list[dict] = []
    offset = 0
    limit = 100

    while True:
        data = client.get_by_href(
            positions_href,
            params={
                "limit": limit,
                "offset": offset,
                "expand": "assortment,slot",
            },
        )

        if isinstance(data, list):
            page = data
        else:
            page = data.get("rows", [])

        if debug and page:
            first = page[0]
            slot_raw = first.get("slot")
            print(
                f"{Fore.YELLOW}[DEBUG] rows={len(page)}"
                f" | первая позиция: {first.get('assortment', {}).get('name', '—')[:40]}"
                f" | slot: {slot_raw}{Style.RESET_ALL}"
            )

        positions.extend(page)

        if len(page) < limit:
            break
        offset += limit

    # Шаг 3: определяем ячейку для каждой позиции
    # Приоритет: slot.name → атрибут товара (fallback)
    product_cell_cache: dict[str, str] = {}  # href → ячейка из атрибутов

    for pos in positions:
        # Способ 1: slot из позиции (место хранения)
        slot: dict | None = pos.get("slot")
        if slot and isinstance(slot, dict) and slot.get("name"):
            pos["_cell"] = slot["name"]
            continue

        # Способ 2: пользовательский атрибут на товаре (fallback)
        assortment: dict = pos.get("assortment") or {}
        if not assortment.get("name"):
            # expand не раскрыл assortment — отдельный запрос
            pos_href = pos.get("meta", {}).get("href", "")
            if pos_href:
                try:
                    full = client.get_by_href(pos_href, params={"expand": "assortment,slot"})
                    pos["assortment"] = full.get("assortment", {})
                    slot = full.get("slot")
                    if slot and isinstance(slot, dict) and slot.get("name"):
                        pos["_cell"] = slot["name"]
                        continue
                    assortment = pos["assortment"]
                except RuntimeError as exc:
                    print(f"{Fore.RED}[WARN] Не удалось получить позицию: {exc}{Style.RESET_ALL}")

        product_href = _product_href(assortment)
        if product_href not in product_cell_cache:
            product_cell_cache[product_href] = _extract_cell_from_attr(
                client, assortment, product_href, cell_attr, debug=debug
            )
        pos["_cell"] = product_cell_cache[product_href]

    return positions


def _product_href(assortment: dict) -> str:
    """Вернуть href товара (для варианта — href родительского товара)."""
    meta = assortment.get("meta", {})
    entity_type = meta.get("type", "")

    # Вариант → берём href родительского product
    if entity_type == "variant":
        product = assortment.get("product")
        if product:
            return product.get("meta", {}).get("href", "")

    return meta.get("href", "")


def _extract_cell_from_attr(
    client: MoySkladClient,
    assortment: dict,
    product_href: str,
    cell_attr: str,
    debug: bool = False,
) -> str:
    """Fallback: поиск ячейки в пользовательских атрибутах товара."""
    # Атрибуты могут присутствовать сразу (если API вернул их)
    cell = _find_attr_value(assortment.get("attributes", []), cell_attr)
    if cell:
        return cell

    # Атрибуты не возвращаются через expand → запрашиваем товар отдельно
    if product_href:
        try:
            product_data = client.get_by_href(product_href)
            if debug:
                attrs = product_data.get("attributes", [])
                print(
                    f"{Fore.YELLOW}[DEBUG] Атрибуты товара ({product_href.split('/')[-1]}): "
                    f"{[a.get('name') for a in attrs]}{Style.RESET_ALL}"
                )
            cell = _find_attr_value(product_data.get("attributes", []), cell_attr)
        except RuntimeError as exc:
            print(f"{Fore.RED}[WARN] Не удалось получить атрибуты товара: {exc}{Style.RESET_ALL}")

    return cell


def _find_attr_value(attributes: list[dict], attr_name: str) -> str:
    """Найти значение атрибута по имени."""
    for attr in attributes:
        if attr.get("name", "").lower() == attr_name.lower():
            raw = attr.get("value", "")
            return str(raw) if raw is not None else ""
    return ""


def sort_key(cell: str) -> tuple:
    """
    Ключ натуральной сортировки для ячеек вида 'A-1-2-3', 'E-6-2-1'.

    Буквенные части сортируются лексикографически,
    числовые — численно. Пустые ячейки — в конец.
    """
    if not cell:
        return (chr(0x10FFFF),)  # пустые → в самый конец

    parts = cell.split("-")
    result: list[tuple] = []
    for part in parts:
        try:
            result.append((0, int(part), ""))
        except ValueError:
            result.append((1, 0, part.upper()))
    return tuple(result)  # type: ignore[return-value]


def apply_sort_to_demand(
    client: MoySkladClient,
    demand_id: str,
    sorted_positions: list[dict],
    debug: bool = False,
) -> None:
    """
    Пересортировать позиции отгрузки через batch-DELETE + POST bulk.

    PUT /entity/demand/{id} игнорирует порядок positions. Единственный
    рабочий способ:
      1. POST /entity/demand/{id}/positions/delete  — batch-удаление (1 запрос)
      2. POST  /entity/demand/{id}/positions        — создать в нужном порядке
    """
    total = len(sorted_positions)

    # Шаг 1: batch-удаляем все текущие позиции одним запросом
    delete_payload = [{"meta": pos["meta"]} for pos in sorted_positions]
    if debug:
        import json as _json_d
        print(
            f"  [DEBUG] POST /positions/delete — {total} позиций.\n"
            f"  Первый элемент: {_json_d.dumps(delete_payload[0], ensure_ascii=False)}"
        )
    print(f"  Удаление {total} позиций (batch)…", end="", flush=True)
    client.post(f"/entity/demand/{demand_id}/positions/delete", delete_payload)
    print(" ✓")

    # Шаг 2: создаём все позиции bulk-запросом в нужном порядке
    create_payload: list[dict] = []
    for pos in sorted_positions:
        entry: dict = {
            "assortment": {"meta": pos["assortment"]["meta"]},
            "quantity": pos["quantity"],
            "price": pos.get("price", 0),
        }
        for field in (
            "discount", "vat", "vatEnabled",
            "overhead", "cost", "trackingCodes",
            "things", "gtd", "country",
        ):
            val = pos.get(field)
            if val is not None:
                entry[field] = val
        slot = pos.get("slot")
        if slot and isinstance(slot, dict) and slot.get("meta"):
            entry["slot"] = {"meta": slot["meta"]}
        create_payload.append(entry)

    if debug:
        import json as _json
        print(
            f"  [DEBUG] POST {len(create_payload)} позиций. Первая:\n"
            f"{_json.dumps(create_payload[0], ensure_ascii=False, indent=2)}"
        )

    print(f"  Создание {len(create_payload)} позиций в новом порядке…")
    client.post(f"/entity/demand/{demand_id}/positions", create_payload)


# ── Verify ────────────────────────────────────────────────────────────────────────────


def _make_snapshot(positions: list[dict]) -> dict:
    """Слепок позиций: кол-во строк, итоговое qty, сумма, Counter и порядок ключей."""
    total_qty = 0.0
    total_sum = 0.0
    items: collections.Counter = collections.Counter()
    order_keys: list[tuple] = []

    for pos in positions:
        qty = float(pos.get("quantity", 0))
        price = float(pos.get("price", 0))
        href = (pos.get("assortment") or {}).get("meta", {}).get("href", "?")
        cell = pos.get("_cell", "")

        total_qty += qty
        total_sum += price * qty
        items[(href, cell)] += qty
        order_keys.append((href, cell))

    return {
        "count": len(positions),
        "total_qty": total_qty,
        "total_sum": total_sum,
        "items": items,
        "order_keys": order_keys,
    }


def _print_verify(before: dict, after: dict) -> bool:
    """Сравнить снапшоты до и после. Возвращает True если всё совпало."""
    count_ok = before["count"] == after["count"]
    qty_ok = abs(before["total_qty"] - after["total_qty"]) < 0.001
    # price в МойСклад хранится в денежных единицах (сотые копейки)
    sum_ok = abs(before["total_sum"] - after["total_sum"]) < 1.0

    def _c(ok: bool) -> str:
        return Fore.GREEN if ok else Fore.RED

    def _i(ok: bool) -> str:
        return "✓" if ok else "✗"

    # Перевод суммы из ден. единиц в рубли (делим на 100 центов и на 100)
    b_rub = before["total_sum"] / 10_000
    a_rub = after["total_sum"] / 10_000

    print(f"\n  {'\u041fараметр':<24} {'\u0414о':>14} {'\u041fосле':>14}")
    print(f"  {'─' * 56}")
    print(
        f"  {'\u0421трок позиций':<24}"
        f" {before['count']:>14} {after['count']:>14}"
        f"  {_c(count_ok)}{_i(count_ok)}{Style.RESET_ALL}"
    )
    print(
        f"  {'\u0418того единиц (шт)':<24}"
        f" {before['total_qty']:>14.0f} {after['total_qty']:>14.0f}"
        f"  {_c(qty_ok)}{_i(qty_ok)}{Style.RESET_ALL}"
    )
    print(
        f"  {'\u0421умма (руб.)':<24}"
        f" {b_rub:>14.2f} {a_rub:>14.2f}"
        f"  {_c(sum_ok)}{_i(sum_ok)}{Style.RESET_ALL}"
    )

    # Позицийный дифф: сравниваем Counterы
    missing: list = []
    extra: list = []
    changed: list = []
    for key in set(before["items"]) | set(after["items"]):
        b_qty = before["items"].get(key, 0.0)
        a_qty = after["items"].get(key, 0.0)
        if abs(b_qty - a_qty) >= 0.001:
            href, cell = key
            sku = href.rsplit("/", 1)[-1][:12]
            label = f"ячейка={cell or '—':8} артикуль=…{sku}"
            if b_qty > 0 and a_qty == 0:
                missing.append((label, b_qty))
            elif b_qty == 0 and a_qty > 0:
                extra.append((label, a_qty))
            else:
                changed.append((label, b_qty, a_qty))

    if missing or extra or changed:
        print(f"\n  {Fore.RED}Расхождения по позициям:{Style.RESET_ALL}")
        for label, qty in missing:
            print(f"  {Fore.RED}\u2717 ПОТЕРЯНО   {label}  qty={qty:.0f}{Style.RESET_ALL}")
        for label, qty in extra:
            print(f"  {Fore.YELLOW}\u2717 ЛИШНЕЕ    {label}  qty={qty:.0f}{Style.RESET_ALL}")
        for label, b, a in changed:
            print(f"  {Fore.YELLOW}\u2717 ИЗМЕНЕНО  {label}  qty={b:.0f}→{a:.0f}{Style.RESET_ALL}")

    # Проверка порядка
    exp = before["order_keys"]
    got = after["order_keys"]
    order_ok = (exp == got)
    order_label = "Порядок позиций"
    print(
        f"  {order_label:<24}"
        f" {'(отправлено)':>14} {'(из API)':>14}"
        f"  {_c(order_ok)}{_i(order_ok)}{Style.RESET_ALL}"
    )
    if not order_ok:
        # Показываем первые расхождения (макс 5)
        diffs = [
            i for i, (e, g) in enumerate(zip(exp, got), 1) if e != g
        ]
        extra_count = len(exp) - len(got) if len(exp) != len(got) else 0
        shown = diffs[:5]
        print(f"  {Fore.RED}  Позиции не в том порядке. Первые расхождения: строки {shown}{Style.RESET_ALL}")
        if extra_count:
            print(f"  {Fore.RED}  Длины не совпадают: отправлено {len(exp)}, получено {len(got)}{Style.RESET_ALL}")

    return count_ok and qty_ok and sum_ok and order_ok and not (missing or extra or changed)


# ── XLSX Export ───────────────────────────────────────────────────

_OUTPUT_DIR = pathlib.Path("output")
_BACKUPS_DIR = _OUTPUT_DIR / "backups"


def _sanitize_filename(name: str) -> str:
    """Убираем символы, запрещённые в именах файлов Windows/Linux."""
    return re.sub(r'[\\/:*?"<>|]', "_", name).strip()


def _extract_barcodes(assortment: dict) -> tuple[str, str]:
    """Вернуть (code128, ean13) из поля barcodes ассортимента."""
    code128 = ""
    ean13 = ""
    for entry in assortment.get("barcodes") or []:
        if not isinstance(entry, dict):
            continue
        if "code128" in entry:
            code128 = str(entry["code128"])
        elif "ean13" in entry:
            ean13 = str(entry["ean13"])
    return code128, ean13


def save_xlsx(sorted_positions: list[dict], demand_name: str) -> pathlib.Path:
    """
    Сохранить отсортированные позиции в xlsx.

    Колонки: A=Артикул, B=EAN13, C=Кол-во, D=Ячейка, E=№ Коробки (пустая)
    Позиции НЕ суммируются — каждая строка == одна позиция.
    """
    _OUTPUT_DIR.mkdir(exist_ok=True)
    safe_name = _sanitize_filename(demand_name)
    out_path = _OUTPUT_DIR / f"{safe_name}.xlsx"

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Позиции"

    # Заголовки
    headers = ["Артикул", "EAN13", "Кол-во", "Ячейка", "№ Коробки"]
    header_fill = openpyxl.styles.PatternFill("solid", fgColor="1F4E79")
    header_font = openpyxl.styles.Font(color="FFFFFF", bold=True)
    header_align = openpyxl.styles.Alignment(horizontal="center", vertical="center")

    for col, title in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=title)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_align

    ws.row_dimensions[1].height = 18

    # Стили данных
    left_align = openpyxl.styles.Alignment(horizontal="left", vertical="center")
    center_align = openpyxl.styles.Alignment(horizontal="center", vertical="center")
    box_font = openpyxl.styles.Font(size=20)
    # Фиксированная высота — не даём Excel расширять строки под шрифт 20pt
    DATA_ROW_HEIGHT = 15.0

    # Строки — одна позиция == одна строка
    for row_idx, pos in enumerate(sorted_positions, 2):
        assortment = pos.get("assortment") or {}
        code = assortment.get("code", "")
        _, ean13 = _extract_barcodes(assortment)
        qty = pos.get("quantity", 0)
        cell_name = pos.get("_cell", "")

        ws.cell(row=row_idx, column=1, value=code)                                   # A: Артикул
        ws.cell(row=row_idx, column=2, value=ean13)                                  # B: EAN13

        qty_cell = ws.cell(row=row_idx, column=3,                                    # C: Кол-во
                           value=int(qty) if float(qty) == int(qty) else qty)
        qty_cell.alignment = left_align

        cell_cell = ws.cell(row=row_idx, column=4, value=cell_name)                  # D: Ячейка
        cell_cell.alignment = left_align

        box_cell = ws.cell(row=row_idx, column=5, value="")                          # E: № Коробки (пустая)
        box_cell.alignment = center_align
        box_cell.font = box_font

        ws.row_dimensions[row_idx].height = DATA_ROW_HEIGHT

    # автоширина колонок — по максимуму контента в каждом столбце
    col_letters = ["A", "B", "C", "D", "E"]
    for col_idx, letter in enumerate(col_letters, 1):
        max_len = 0
        for row in ws.iter_rows(min_col=col_idx, max_col=col_idx):
            for cell in row:
                if cell.value is not None:
                    max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[letter].width = max(max_len + 3, 12)  # минимум 12

    wb.save(out_path)
    return out_path


# ── Backup ───────────────────────────────────────────────────────────────────


def save_backup(
    positions: list[dict],
    demand_id: str,
    demand_name: str,
) -> pathlib.Path:
    """Сохранить бэкап позиций отгрузки.

    Формат файла:
      строка 1  — название документа (для человека)
      строки 2+ — JSON-данные позиций
    """
    backup_dir = _BACKUPS_DIR / demand_id
    backup_dir.mkdir(parents=True, exist_ok=True)

    ts = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
    safe_name = _sanitize_filename(demand_name)
    out_path = backup_dir / f"{ts}_{safe_name}.json"

    payload: list[dict] = []
    for pos in positions:
        assortment_meta = (pos.get("assortment") or {}).get("meta")
        if not assortment_meta:
            print(
                f"{Fore.YELLOW}[WARN] save_backup: пропущена позиция без meta ассортимента.{Style.RESET_ALL}"
            )
            continue
        entry: dict = {
            "assortment": {"meta": assortment_meta},
            "quantity": pos["quantity"],
            "price": pos.get("price", 0),
        }
        for field in (
            "discount", "vat", "vatEnabled",
            "overhead", "cost", "trackingCodes",
            "things", "gtd", "country",
        ):
            val = pos.get(field)
            if val is not None:
                entry[field] = val
        slot = pos.get("slot")
        if slot and isinstance(slot, dict) and slot.get("meta"):
            entry["slot"] = {"meta": slot["meta"]}
        payload.append(entry)

    with out_path.open("w", encoding="utf-8") as f:
        f.write(demand_name + "\n")
        json.dump(payload, f, ensure_ascii=False, indent=2)

    return out_path


def list_backups(demand_id: str) -> list[dict]:
    """Вернуть список бэкапов для конкретной отгрузки (новые — первыми).

    Каждый элемент: {demand_name, file, ts_display}.
    """
    backup_dir = _BACKUPS_DIR / demand_id
    if not backup_dir.exists():
        return []

    backups: list[dict] = []
    for bfile in sorted(backup_dir.glob("*.json"), reverse=True):
        try:
            with bfile.open("r", encoding="utf-8") as f:
                demand_name = f.readline().rstrip("\n")
        except OSError:
            demand_name = "?"
        # Имя файла: 20260305_142310_ОТ-12345.json → показываем дату+время
        parts = bfile.stem.split("_", 2)
        ts_display = f"{parts[0][:4]}-{parts[0][4:6]}-{parts[0][6:]} {parts[1][:2]}:{parts[1][2:4]}:{parts[1][4:]}" if len(parts) >= 2 else bfile.stem
        backups.append({"demand_name": demand_name, "file": bfile, "ts_display": ts_display})
    return backups


def load_backup(bfile: "pathlib.Path") -> list[dict]:
    """Прочитать бэкап: пропустить первую строку (имя), остальное — JSON."""
    with bfile.open("r", encoding="utf-8") as f:
        f.readline()  # строка с именем документа
        return json.load(f)


def restore_demand_from_backup(
    client: "MoySkladClient",
    demand_id: str,
    backup_positions: list[dict],
) -> None:
    """Восстановить позиции отгрузки из бэкапа.

    1. Получить текущие позиции (для batch-DELETE)
    2. Удалить все текущие позиции
    3. Создать позиции из бэкапа в исходном порядке
    """
    print("  Получение текущих позиций…", end="", flush=True)
    demand_data = client.get(f"/entity/demand/{demand_id}")
    positions_href = demand_data.get("positions", {}).get("meta", {}).get("href", "")

    current_positions: list[dict] = []
    if positions_href:
        offset, limit = 0, 100
        while True:
            data = client.get_by_href(positions_href, params={"limit": limit, "offset": offset})
            page = data.get("rows", []) if isinstance(data, dict) else data
            current_positions.extend(page)
            if len(page) < limit:
                break
            offset += limit
    print(f" {len(current_positions)} шт.")

    if current_positions:
        delete_payload = [{"meta": p["meta"]} for p in current_positions]
        print(f"  Удаление {len(current_positions)} позиций…", end="", flush=True)
        client.post(f"/entity/demand/{demand_id}/positions/delete", delete_payload)
        print(" ✓")

    print(f"  Восстановление {len(backup_positions)} позиций из бэкапа…", end="", flush=True)
    client.post(f"/entity/demand/{demand_id}/positions", backup_positions)
    print(" ✓")


# ── QR Terminal ──────────────────────────────────────────────────────────────


def print_qr_terminal(demand_name: str) -> None:
    """Вывести QR-код с номером отгрузки прямо в терминал для сканирования на ТСД."""
    qr = _qrcode.QRCode(
        version=None,
        error_correction=_qrcode.constants.ERROR_CORRECT_M,
        box_size=1,
        border=2,
    )
    qr.add_data(demand_name)
    qr.make(fit=True)

    label = "QR-код отгрузки"
    # inner — ширина содержимого между ╔ и ╗, динамически под имя отгрузки
    inner = max(len(label) + 4, len(demand_name) + 4, 30)
    top    = f"╔═ {label} {'═' * (inner - len(label) - 3)}╗"
    middle = f"║ {demand_name:<{inner - 1}}║"
    bottom = f"╚{'═' * inner}╝"

    print(f"\n{Fore.CYAN}{top}{Style.RESET_ALL}")
    print(f"{Fore.CYAN}{middle}{Style.RESET_ALL}")
    print(f"{Fore.CYAN}{bottom}{Style.RESET_ALL}")
    qr.print_ascii(invert=True)


# ── Display ───────────────────────────────────────────────────────────────────


def print_demands_table(demands: list[dict]) -> None:
    # Собираем строки
    rows = []
    for i, d in enumerate(demands, 1):
        rows.append((
            str(i),
            d.get("moment", "")[:10],
            d.get("name", "—"),
            (d.get("agent") or {}).get("name", "—"),
            (d.get("state") or {}).get("name", "—"),
        ))

    # Динамическая ширина колонок (с ограничением максимума)
    wn  = max(len("№"),          max(len(r[0]) for r in rows))
    wd  = max(len("Дата"),       max(len(r[1]) for r in rows))
    wno = min(max(len("Номер"),  max(len(r[2]) for r in rows)), 18)
    wa  = min(max(len("Контрагент"), max(len(r[3]) for r in rows)), 38)
    ws  = max(len("Статус"),     max(len(r[4]) for r in rows))

    sep = "  "
    total = wn + wd + wno + wa + ws + len(sep) * 4

    header = (
        f"{'№':>{wn}}{sep}{'Дата':<{wd}}{sep}{'Номер':<{wno}}{sep}"
        f"{'Контрагент':<{wa}}{sep}{'Статус':<{ws}}"
    )
    print(f"\n{Fore.CYAN}{header}{Style.RESET_ALL}")
    print("─" * total)

    for num, date, name, agent, state in rows:
        print(
            f"{Fore.WHITE}{num:>{wn}}{Style.RESET_ALL}{sep}"
            f"{date:<{wd}}{sep}"
            f"{name[:wno]:<{wno}}{sep}"
            f"{agent[:wa]:<{wa}}{sep}"
            f"{state}"
        )


def print_positions_table(sorted_positions: list[dict]) -> None:
    # Собираем строки
    rows = []
    for i, pos in enumerate(sorted_positions, 1):
        name = (pos.get("assortment") or {}).get("name", "—")
        qty  = pos.get("quantity", 0)
        cell = pos.get("_cell", "")
        rows.append((str(i), cell, name, qty))

    # Динамическая ширина
    wn  = max(len("№"),           max(len(r[0]) for r in rows))
    wc  = min(max(len("Ячейка"),  max(len(r[1]) for r in rows)), 16)
    wna = min(max(len("Наименование"), max(len(r[2]) for r in rows)), 52)
    wq  = max(len("Кол-во"),      max(len(f"{r[3]:.0f}") for r in rows))

    sep   = "  "
    total = wn + wc + wna + wq + len(sep) * 3

    header = (
        f"{'№':>{wn}}{sep}{'Ячейка':<{wc}}{sep}"
        f"{'Наименование':<{wna}}{sep}{'Кол-во':>{wq}}"
    )
    print(f"\n{Fore.CYAN}{header}{Style.RESET_ALL}")
    print("─" * total)

    for num, cell, name, qty in rows:
        cell_pad = f"{cell[:wc]:<{wc}}"
        if cell:
            cell_str = f"{Fore.YELLOW}{cell_pad}{Style.RESET_ALL}"
        else:
            cell_str = f"{Fore.RED}{'—':<{wc}}{Style.RESET_ALL}"

        print(
            f"{num:>{wn}}{sep}"
            f"{cell_str}{sep}"
            f"{name[:wna]:<{wna}}{sep}"
            f"{qty:>{wq}.0f}"
        )


# ── Helpers ───────────────────────────────────────────────────────────────────


def _err(msg: str, exit_code: int = 1) -> None:
    print(f"{Fore.RED}{msg}{Style.RESET_ALL}", file=sys.stderr)
    sys.exit(exit_code)


def _demand_id_from(demand: dict) -> str:
    """Извлечь UUID отгрузки. Берём поле id напрямую — безопаснее, чем парсить href."""
    return demand["id"]


def _pick_demand(demands: list[dict]) -> dict:
    """Интерактивный выбор отгрузки по номеру."""
    total = len(demands)
    while True:
        try:
            raw = input(
                f"\n{Fore.CYAN}Введите номер отгрузки "
                f"[1–{total}] (0 — выход): {Style.RESET_ALL}"
            ).strip()
            num = int(raw)
        except (ValueError, EOFError):
            print(f"{Fore.RED}Введите целое число.{Style.RESET_ALL}")
            continue

        if num == 0:
            print("Выход.")
            sys.exit(0)
        if 1 <= num <= total:
            return demands[num - 1]
        print(f"{Fore.RED}Введите число от 1 до {total}.{Style.RESET_ALL}")


# ── Entry Point ───────────────────────────────────────────────────────────────


def main(config: AppConfig) -> None:  # noqa: D401
    token = load_token()
    client = MoySkladClient(token)

    # 1. Найти статус
    print(f"{Fore.CYAN}Поиск статуса «{config.state_name}»…{Style.RESET_ALL}")
    state_href = find_state_href(client, config.state_name)
    print(f"{Fore.GREEN}✓ Статус найден{Style.RESET_ALL}")

    # 2. Загрузить отгрузки
    since = datetime.datetime.now() - datetime.timedelta(days=config.days)
    print(
        f"Загрузка отгрузок за последние {config.days} дн. "
        f"(с {since.strftime('%Y-%m-%d')})…"
    )
    demands = fetch_demands(client, since, state_href)

    if not demands:
        print(f"{Fore.YELLOW}Отгрузок не найдено.{Style.RESET_ALL}")
        return

    print(f"{Fore.GREEN}✓ Найдено: {len(demands)}{Style.RESET_ALL}")

    # 3. Показать таблицу и получить выбор пользователя
    print_demands_table(demands)
    selected = _pick_demand(demands)
    demand_id = _demand_id_from(selected)
    demand_name = selected.get("name", demand_id)

    # 4. Загрузить позиции и определить ячейки
    print(
        f"\nЗагрузка позиций «{Fore.CYAN}{demand_name}{Style.RESET_ALL}»…"
    )
    positions = fetch_positions(client, demand_id, config.cell_attr, debug=config.debug)

    if not positions:
        print(f"{Fore.YELLOW}В отгрузке нет позиций.{Style.RESET_ALL}")
        return

    without_cell = sum(1 for p in positions if not p.get("_cell"))
    if without_cell:
        print(
            f"{Fore.YELLOW}⚠ {without_cell} из {len(positions)} позиций "
            f"без атрибута «{config.cell_attr}» — будут в конце списка.{Style.RESET_ALL}"
        )

    # 5. Отсортировать и отобразить
    sorted_positions = sorted(
        positions, key=lambda p: sort_key(p.get("_cell", ""))
    )
    print_positions_table(sorted_positions)

    # 6. Меню действий
    # --config.apply = автовыбор «2» (сортировка + документ)
    if config.apply:
        action = 2
    else:
        available_backups = list_backups(demand_id)
        restore_line = (
            f"  {Fore.WHITE}9{Style.RESET_ALL} — восстановить из бэкапа"
            f" {Fore.YELLOW}({len(available_backups)} доступно){Style.RESET_ALL}\n"
            if available_backups else ""
        )
        valid_actions = {0, 1, 2, 3} | ({9} if available_backups else set())
        range_hint = "0–3/9" if available_backups else "0–3"
        print(
            f"\n{Fore.CYAN}Выберите действие:{Style.RESET_ALL}\n"
            f"  {Fore.WHITE}1{Style.RESET_ALL} — применить сортировку в МойСклад\n"
            f"  {Fore.WHITE}2{Style.RESET_ALL} — применить сортировку + сохранить xlsx\n"
            f"  {Fore.WHITE}3{Style.RESET_ALL} — только сохранить xlsx (без изменений в МойСклад)\n"
            f"{restore_line}"
            f"  {Fore.WHITE}0{Style.RESET_ALL} — отмена"
        )
        action = 0
        while True:
            try:
                raw = input(
                    f"{Fore.YELLOW}Ваш выбор [{range_hint}]: {Style.RESET_ALL}"
                ).strip()
                action = int(raw)
            except (ValueError, EOFError):
                print(f"{Fore.RED}Введите корректный номер.{Style.RESET_ALL}")
                continue
            if action in valid_actions:
                break
            print(f"{Fore.RED}Недопустимый выбор.{Style.RESET_ALL}")

    if action == 0:
        print("Отменено.")

    # Бэкап — сразу после выбора действия, до любой обработки
    if action in (1, 2, 3):
        backup_path = save_backup(sorted_positions, demand_id, demand_name)
        print(f"{Fore.GREEN}✓ Бэкап сохранён: {backup_path}{Style.RESET_ALL}")

    # Применить сортировку в МойСклад
    if action in (1, 2):
        before_snap = _make_snapshot(sorted_positions)
        print("Применяю сортировку…")
        apply_sort_to_demand(client, demand_id, sorted_positions, debug=config.debug)

        print("  Сверка данных…")
        after_positions = fetch_positions(
            client, demand_id, config.cell_attr, debug=False
        )
        after_snap = _make_snapshot(after_positions)
        ok = _print_verify(before_snap, after_snap)

        if ok:
            print(f"\n{Fore.GREEN}✓ Порядок обновлён. Данные совпали.{Style.RESET_ALL}")
        else:
            print(f"\n{Fore.RED}✗ ВНИМАНИЕ: несоответствие данных после сохранения!{Style.RESET_ALL}")

    # Сохранить xlsx
    if action in (2, 3):
        xlsx_path = save_xlsx(sorted_positions, demand_name)
        print(f"{Fore.GREEN}✓ Сохранено: {xlsx_path}{Style.RESET_ALL}")

    # QR — для всех действий, создающих/обновляющих документ (1, 2, 3)
    if action in (1, 2, 3):
        print_qr_terminal(demand_name)

    # Восстановить из бэкапа
    if action == 9:
        backups = list_backups(demand_id)
        print(f"\n{Fore.CYAN}Доступные бэкапы для «{demand_name}»:{Style.RESET_ALL}")
        for i, b in enumerate(backups, 1):
            print(
                f"  {Fore.WHITE}{i}{Style.RESET_ALL}"
                f" — {Fore.YELLOW}{b['ts_display']}{Style.RESET_ALL}"
                f"  {b['file'].name}"
            )
        while True:
            try:
                raw = input(
                    f"{Fore.YELLOW}Выберите бэкап [1–{len(backups)}] (0 — отмена): {Style.RESET_ALL}"
                ).strip()
                num = int(raw)
            except (ValueError, EOFError):
                print(f"{Fore.RED}Введите целое число.{Style.RESET_ALL}")
                continue
            if num == 0:
                print("Отменено.")
                break
            if 1 <= num <= len(backups):
                chosen = backups[num - 1]
                backup_positions = load_backup(chosen["file"])
                print(
                    f"\nВосстановление «{demand_name}» "
                    f"из бэкапа {chosen['ts_display']}…"
                )
                restore_demand_from_backup(client, demand_id, backup_positions)
                print(f"{Fore.GREEN}✓ Отгрузка восстановлена.{Style.RESET_ALL}")
                break
            print(f"{Fore.RED}Введите число от 1 до {len(backups)}.{Style.RESET_ALL}")

    # Меню смены статуса перед выходом
    if not config.apply and action in (1, 2, 3):
        assembled_href = find_state_href(client, "Собран")
        assembly_href = find_state_href(client, config.state_name)

        while True:
            print(
                f"\n{Fore.CYAN}Выберите действие:{Style.RESET_ALL}\n"
                f"  {Fore.WHITE}1{Style.RESET_ALL} — сменить статус на {Fore.GREEN}'Собран'{Style.RESET_ALL}\n"
                f"  {Fore.WHITE}0{Style.RESET_ALL} — выход"
            )
            try:
                raw = input(f"{Fore.YELLOW}Ваш выбор [0–1]: {Style.RESET_ALL}").strip()
                choice = int(raw)
            except (ValueError, EOFError):
                print(f"{Fore.RED}Введите 0 или 1.{Style.RESET_ALL}")
                continue

            if choice == 0:
                break
            if choice == 1:
                print(f"Обновляю статус «{demand_name}» → Собран…", end="", flush=True)
                change_demand_state(client, demand_id, assembled_href)
                print(f" {Fore.GREEN}✓{Style.RESET_ALL}")

                # После успешной смены — предложить откат назад
                while True:
                    print(
                        f"\n{Fore.CYAN}Выберите действие:{Style.RESET_ALL}\n"
                        f"  {Fore.WHITE}1{Style.RESET_ALL} — вернуть статус на {Fore.YELLOW}'{config.state_name}'{Style.RESET_ALL}\n"
                        f"  {Fore.WHITE}0{Style.RESET_ALL} — выход"
                    )
                    try:
                        raw2 = input(f"{Fore.YELLOW}Ваш выбор [0–1]: {Style.RESET_ALL}").strip()
                        choice2 = int(raw2)
                    except (ValueError, EOFError):
                        print(f"{Fore.RED}Введите 0 или 1.{Style.RESET_ALL}")
                        continue
                    if choice2 == 0:
                        break
                    if choice2 == 1:
                        print(f"Возвращаю статус «{demand_name}» → {config.state_name}…", end="", flush=True)
                        change_demand_state(client, demand_id, assembly_href)
                        print(f" {Fore.GREEN}✓{Style.RESET_ALL}")
                        break
                break  # выход из внешнего цикла
            print(f"{Fore.RED}Недопустимый выбор.{Style.RESET_ALL}")


if __name__ == "__main__":
    tyro.cli(main)
